import openpyxl
import pandas
import xlrd
from xlrd.sheet import ctype_text


class ExcelUtil(object):

    @staticmethod
    def format_number(number):
        """
            删除小数点后多余的0
        """
        if isinstance(number, int):
            return number
        if isinstance(number, float):
            # 删除小数点后多余的0
            number = str(number).rstrip('0')

            # 只剩小数点直接转int，否则转回float
            if number.endswith('.'):
                number = int(number.rstrip('.'))
            else:
                number = float(number)
            return number

    @staticmethod
    def read_excel_xlrd(excel_file_path, field_list, sheet_idx_list=None, sheet_name=None):
        """
            excel_file_path: 文件路径
            field_list: 字段列表
            sheet_idx_list: 页号，第一页是0
            sheet_name: 页名
        """
        book = xlrd.open_workbook(excel_file_path)
        sheet_total = book.nsheets
        sheet_idx_max = sheet_total - 1

        # get target sheet idx list
        if sheet_idx_list and type(sheet_idx_list) == int and sheet_idx_list >= 0:
            sheet_idx_list = [sheet_idx_list]
        if sheet_idx_list and type(sheet_idx_list) == list and len(sheet_idx_list) > 0:
            print(">> sheet_idx_list: {}".format(sheet_idx_list))
        else:
            sheet_idx_list = []
            for i in range(sheet_total):
                sheet_idx_list.append(i)

        # get file content
        file_content = []
        for idx_sheet in sheet_idx_list:
            if idx_sheet > sheet_idx_max:
                continue

            sheet = book.sheet_by_index(idx_sheet)
            sheet_name_current = str(sheet.name).lower().strip()
            row_total = sheet.nrows
            if sheet_name and str(sheet_name).lower().strip() != sheet_name_current:
                # 指定页名时，只匹配指定页的数据
                continue
            for idx_row in range(row_total):
                row = sheet.row(idx_row)
                col_total = len(row)

                line_content = []
                for idx_col in range(col_total):
                    col = row[idx_col]
                    col_ctype = col.ctype
                    col_ctype_name = ctype_text[col_ctype]
                    col_value = col.value

                    value = col_value
                    if col_ctype_name == "xldate":
                        value = xlrd.xldate.xldate_as_datetime(value, 0)
                    elif col_ctype_name == "number":
                        value = ExcelUtil.format_number(value)
                    line_content.append(value)
                line_dict = dict(zip(field_list, line_content))
                file_content.append(line_dict)
            # end for
        # end for
        return file_content

    @staticmethod
    def read_excel_pds(excel_file_path, sheet_name, field_list=None):
        """
            首行不是标题行的需要特殊处理
            返回内容不包含首行
        :param excel_file_path:
        :param sheet_name:
        :param field_list:
        :return:
        """
        file_content = []
        
        df = pandas.read_excel(excel_file_path, sheet_name=sheet_name)
        values = df.values
        if field_list is None:
            return values
        
        for i in range(0, len(values)):
            line_dict = dict(zip(field_list, values[i]))
            file_content.append(line_dict)
        # end for
        
        return file_content

    @staticmethod
    def read_excel(excel_file_path, sheet_name):
        workbook = openpyxl.load_workbook(excel_file_path, read_only=False)
        sheet_names = workbook.sheetnames
        if sheet_name not in sheet_names:
            raise UserWarning("{}不存在！".format(sheet_name))
    
        sheet = workbook[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column
    
        file_content = []
        for i in range(0, max_row):
            file_line = []
            for j in range(0, max_col):
                value = sheet.cell(row=i + 1, column=j + 1).value
                file_line.append(value)
            file_content.append(file_line)
        # end for
        workbook.close()
    
        return file_content
    
    @staticmethod
    def write_excel(data_list, excel_file_path, sheet_title=''):
        if data_list is None or len(data_list) == 0:
            print("no data to write!")
            return
        else:
            print("data_size: {}".format(len(data_list)))
        workbook = openpyxl.Workbook()
        if sheet_title is None or sheet_title == '':
            sheet = workbook.active
            sheet.title = 'Sheet1'
        else:
            sheet = workbook.create_sheet(title=sheet_title)
        
        # 标题行
        field_list = []
        first_line = data_list[0]
        if type(first_line) != dict:
            first_line = first_line.__dict__
        for key in first_line:
            field_list.append(key)
        for i in range(0, len(field_list)):
            sheet.cell(row=1, column=(i + 1), value=field_list[i])
        
        # 内容
        idx_row = 2
        for line in data_list:
            if type(line) != dict:
                line = line.__dict__
            for i in range(0, len(field_list)):
                value = line[field_list[i]]
                try:
                    sheet.cell(row=idx_row, column=(i + 1), value=value)
                except Exception as err:
                    print(i+1, field_list[i], value)
                    raise err
            idx_row += 1
        # end for

        print(excel_file_path)
        workbook.save(excel_file_path)

    @staticmethod
    def write_excel_multi_sheet(data_dict, excel_file_path):
        """
            多页写入excel
        :param data_dict: ['sheet1': [data], 'sheet1': [data2]]
        :param excel_file_path:
        :return:
        """
        workbook = openpyxl.Workbook()

        idx = 0
        for sheet_name in data_dict:
            idx += 1
            data_list = data_dict[sheet_name]
            if data_list is None or len(data_list) == 0:
                print(">> no data to write! {}".format(sheet_name))
                continue
            else:
                print(">> data_size: {}".format(len(data_list)))

            if idx == 1:
                sheet = workbook.active
                sheet.title = sheet_name
            else:
                sheet = workbook.create_sheet(title=sheet_name)
    
            # 标题行
            field_list = []
            first_line = data_list[0]
            if type(first_line) != dict:
                first_line = first_line.__dict__
            for key in first_line:
                field_list.append(key)
            for i in range(0, len(field_list)):
                sheet.cell(row=1, column=(i + 1), value=field_list[i])
        
            # 内容
            idx_row = 2
            for line in data_list:
                if type(line) != dict:
                    line = line.__dict__
                for i in range(0, len(field_list)):
                    value = ""
                    if field_list[i] in line:
                        value = line[field_list[i]]
                    sheet.cell(row=idx_row, column=(i + 1), value=value)
                idx_row += 1
            # end for
    
        print(excel_file_path)
        workbook.save(excel_file_path)


if __name__ == '__main__':
    pass
